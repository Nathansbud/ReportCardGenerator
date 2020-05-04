import os
import re
from threading import Thread
from enum import Enum
import webbrowser
import json

import googleapiclient.errors
import httplib2
import openpyxl.utils.exceptions
from PyQt5.QtGui import QColor, QStandardItem
from PyQt5.QtWidgets import QLineEdit, QInputDialog, QFileDialog, QTableWidgetItem

from cuter import Button, Label, Dropdown, Textarea, ColorSelector, Checkbox, Table, Multidialog, ColorBox, ColorButton
from cuter import app, window
from grades import GradeSet, GradeType, GradeScheme, GradeScale, load_grades
from preferences import prefs
from sheets import get_sheet, write_sheet, make_report_sheet
from veracross import get_class_json
from util import *

'''
Todo:
    High:
        - Overhaul classes so they can be edited in-program (load entire sheet in-program, use interface of Sheet Creator)
        - Async spreadsheet loading (modal dialog & QThread)
        - GUI so that user doesn't have to deal with weird text macros
        - Grade scheme tab rework
    Medium:
        - updateTable desperately needs an overhaul
        - Format student dropdown for unfinished reports
        - Support for multi-paragraph reports
        - Fix multidialogs to be...not a freaking mess (especially initialize settings)
        - UI overhaul (it's friggin ugly)
    Low:
        - Use Sheets API file selector
        - oldText/cell saving, updating; Enter key tab down kinda buggy (hackish fix rn by doing enter behavior in validation functions)
        - setUI and all graphical stuff...is a damn mess
        - Support for RTF (italics)
    Very Low:
        - Rework editOption to use Multidialog
    
Bugs: 
    High:
    Medium:
    Low:
        - Removed columns aren't removed unless sandwiched
        - Dropdown dialogs can occasionally be OK'd without option selected
    Very Low:
        - Sentence dropdowns occasionally aren't updated if removed (causing gap between dropdowns)
'''

excel_extensions = ["xlsx", "xlsm", "xltx", "xltm", "xlw"]
report_sheet = prefs.get_pref("report_sheet")

def file_select(name, extensions):
    dialog = QFileDialog()
    dialog.setFileMode(QFileDialog.AnyFile)
    dialog.setNameFilter(f"{name} ({' '.join([f'*.{e}' for e in extensions])})")
    if dialog.exec():
        return dialog.selectedFiles()[0]
    return False #False = Failed

def directory_select():
    dialog = QFileDialog()
    dialog.setFileMode(QFileDialog.DirectoryOnly)
    if dialog.exec():
        return dialog.selectedFiles()[0]
    return False

report_column = "D"
report_column_index = 3

grades_column = "E"
grades_column_index = 4

row_offset = 2

sheet = []
all_tab_pairs = []
all_tabs = []
class_tabs = []
sentence_tabs = []
class_students = []
preset_list = {}
sentences = []  # Should be populated with SentenceGroup elements
grade_scheme_tabs = []
grade_rules = []
user_templates = {}



class_label = Label("Reports", "Class: ", window.width() / 2.5, 15)
class_dropdown = Dropdown("Reports", class_label.xw(), class_label.y(), [tab for tab in class_tabs])

student_label = Label("Reports", "Name: ", 50, 75)
student_dropdown = Dropdown("Reports", student_label.xw(), student_label.y(), [])
student_dropdown.setObjectName("StudentDropdown")


open_grades_from_reports_button = Button("Reports", "Grades", student_dropdown.xw(), student_dropdown.y(), False)
add_sentence_button = Button("Reports", "Add Sentence", open_grades_from_reports_button.xw(), student_dropdown.y(), False)
reload_all_button = Button("Reports", "Reload", add_sentence_button.xw(), add_sentence_button.y(), False)

launch_report_sheet_button = Button("Reports", "Open Reports", reload_all_button.xw(), reload_all_button.y(), False)
toggle_advanced_button = Button("Reports", "Toggle Advanced", launch_report_sheet_button.xw(), launch_report_sheet_button.y(), False)
help_button = Button("Reports", "Help", toggle_advanced_button.xw(), toggle_advanced_button.y(), False)


preset_dropdown = Dropdown("Reports", student_dropdown.x(), student_dropdown.yh(), [], False)
preset_button = Button("Reports", "Generate (Preset)", preset_dropdown.xw(), preset_dropdown.y(), False)
grade_button = Button("Reports", "Generate (Grades)", preset_button.xw(), preset_dropdown.y(), False)

open_setup_from_report_button = Button("Reports", "Sheet Setup", grade_button.xw(), grade_button.y(), False)

generate_button = Button("Reports", "Generate", window.width()/2 - 20, 410)
report_area = Textarea("Reports", "", 0, 450, window.width(), 250)
submit_button = Button("Reports", "Submit", window.width()/2 - 20, 700)

copy_button = Button("Reports", "Copy Report", submit_button.xw(), submit_button.y(), False)

color_selector = ColorSelector("Reports")

open_preferences_from_reports_button = Button("Reports", "Preferences", 0, 0, False)
open_preferences_from_reports_button.move(window.width() - open_preferences_from_reports_button.width(), 0)

open_reports_from_preferences_button = Button("Preferences", "Open Reports", window.width() - 50, 0, False)
open_reports_from_preferences_button.move(window.width() - open_reports_from_preferences_button.width(), 0)

open_reports_from_grades_button = Button("Grades", "Open Reports", window.width() - 50, 0, False)
open_reports_from_grades_button.move(window.width() - open_reports_from_preferences_button.width(), 0)

open_reports_from_setup_button = Button("Setup", "Back", 0, 0, False, shown=False)

open_grades_from_reports_button.clicked.connect(lambda: window.switchScreen("Grades"))
open_preferences_from_reports_button.clicked.connect(lambda: window.switchScreen("Preferences"))
open_setup_from_report_button.clicked.connect(lambda: window.switchScreen("Setup"))

open_reports_from_grades_button.clicked.connect(lambda: window.switchScreen("Reports"))
open_reports_from_preferences_button.clicked.connect(lambda: window.switchScreen("Reports"))
open_reports_from_setup_button.clicked.connect(lambda: window.switchScreen("Reports"))


background_color_button = Button("Preferences", "BG Color", window.width() - 150, open_reports_from_preferences_button.yh(), False)
text_color_button = Button("Preferences", "Text Color", window.width() - 150, background_color_button.yh(), False)
label_color_button = Button("Preferences", "Label Color", window.width() - 150, text_color_button.yh(), False)

background_color_button.clicked.connect(lambda: color_selector.updateColor("Background Color", "bg_color"))
text_color_button.clicked.connect(lambda: color_selector.updateColor("Text Color", "txt_color"))
label_color_button.clicked.connect(lambda: color_selector.updateColor("Label Color", "lbl_color"))

# UI Cleanup, might bring back
# reload_sentences_button = Button("Reports", "Sentences", reload_all_button.xw(), reload_all_button.y(), False)
# reload_schemes_button = Button("Reports", "Schemes", reload_sentences_button.xw(), reload_sentences_button.y(), False)

def sentence_tab(s=None):
    if not s: s=class_dropdown.currentText()
    return "Sentences " + s.split("-")[0]

class Student:
    pronouns = {
        "M": ["he", "his", "him", "his", "himself"],
        "F": ["she", "her", "her", "hers", "herself"],
        "T": ["they", "their", "them", "theirs", "themselves"]
    }

    def __init__(self, first_name, last_name, gender, report, grades, classroom, offset):
        self.first_name = first_name
        self.last_name = last_name
        self.report = report
        self.grades = grades
        self.gender = gender
        self.classroom = classroom
        self.offset = offset
        if self.report: self.submitted = True
        else: self.submitted = False

    def submit_report(self):
        global report_sheet
        global report_column
        global grades_column
        global grades_column_index
        global sheet
        global student_dropdown
        global class_students

        if not self.report: report = ""
        else:
            self.report = self.report.strip()
            report = self.report

        if prefs.get_pref('is_web'):
            write_sheet(report_sheet, [[report]], r="{}!{}".format(self.classroom, report_column + str(self.offset)))
        else:
            sheet[self.classroom][report_column + str(self.offset)] = report
            sheet.save(prefs.get_pref('report_sheet'))

        if self.report: self.submitted = True
        else: self.submitted = False

    def write_grades(self):
        global sheet
        if prefs.get_pref('is_web'):
            write_sheet(report_sheet, [[g['grade'] for g in self.grades.values()]], r="{}!{}:{}".format(self.classroom, grades_column + str(self.offset), index_to_column(grades_column_index + len(self.grades)) + str(self.offset)))
        else:
            grades = [g['grade'] for g in self.grades.values()]
            for i in range(len(self.grades)):
                sheet[self.classroom][index_to_column(grades_column_index+i)+str(self.offset)] = grades[i] if not None else ""
            sheet.save(prefs.get_pref('report_sheet'))

    def get_pronouns(self): return Student.pronouns[self.gender] if self.gender in Student.pronouns else Student.pronouns["T"]
    def full_name(self): return ((self.first_name if self.first_name else "") + " " + (self.last_name if self.last_name else "")).strip()

def replace_generics(fmt, recurse=False):
    global student_dropdown
    global class_students
    global user_templates

    idx = student_index()
    if idx is not False:
        fmt = make_lowercase_generics(fmt)
        current_student = class_students[student_index()]
        ps = current_student.get_pronouns()

        replace_set = {
            "{name}":current_student.first_name, "@":current_student.first_name,
            "{p1}":ps[0], "#":ps[0],
            "{p2}":ps[1], "$":ps[1],
            "{p3}":ps[2], "%":ps[2],
            "{p4}":ps[3], "^":ps[3],
            "{p5}":ps[4], "`":ps[4]
        }

        if not recurse:
            for t, v in user_templates.items():
                if not 'class' in v or class_dropdown.currentText().startswith(v['class']):
                    tg = replace_generics(v['value'], True)
                    replace_set[t] = tg

        fmt = fmt.split(Preset.prefix['grade'])[0].split(Preset.prefix['linear'])[0].strip()
        for key in replace_set:
            fmt = fmt.replace(key, replace_set[key])
        #Error Check
        for mp, fp, tp in zip(Student.pronouns['M'], Student.pronouns['F'], Student.pronouns['T']):
            #(?i) = case insensitive, \b = word boundary, (?<!\\) = negative look behind to check not starting with backslash
            if current_student.gender == "M":
                fmt = re.sub(f'(?i)\\b(?<!\\\\){fp}\\b', mp, fmt)
            elif current_student.gender == "F":
                fmt = re.sub(f'(?i)\\b(?<!\\\\){mp}\\b', fp, fmt)
            elif current_student.gender == "T":
                fmt = re.sub(f'(?i)\\b(?<!\\\\)({mp}|{fp})\\b', tp, fmt)

        fmt = fmt.replace("\\", "")

        if current_student.gender == "T":
            fmt = fmt.replace("they is", "they are")

        if not recurse:
            punctuationIndices = []
            if len(fmt) > 0: fmt = fmt[0].upper() + fmt[1:]
            for i in range(0, len(fmt)):
                if fmt[i] == "." or fmt[i] == "!" or fmt[i] == "?": punctuationIndices.append(i)

            for index in punctuationIndices:
                if index + 2 < len(fmt):
                    fmt = fmt[0:index + 2] + fmt[index + 2].upper() + fmt[index + 3:]

        return fmt.strip()
    else: return fmt

def make_lowercase_generics(fmt):
    substr = fmt
    has_match = True
    matches = []
    while has_match:
        matched = re.search('{(.*?)}', substr)
        if not matched:
            break
        else:
            matches.append((matched.start() + (len(fmt) - len(substr)), matched.end() + (len(fmt) - len(substr))))
            substr = substr[matched.end():]

    substr = fmt
    for match in matches: substr = substr[0:match[0]] + fmt[match[0]:match[1]].lower() + fmt[match[1]:]
    return substr

class SentenceGroup:
    dialog = QInputDialog(window.getScreen("Reports"))
    dialog.setOption(QInputDialog.UseListViewForComboBoxItems)
    def __init__(self, label, x, y, options, index):
        self.index = index

        self.x = x
        self.y = y

        self.checkbox = Checkbox("Reports", x, y)
        self.label = Label("Reports", label, x + self.checkbox.width(), y)
        self.dropdown = Dropdown("Reports", self.label.xw(), y, [replace_generics(option) for option in options])
        self.dropdown.options = options

        self.add = Button("Reports", "+", self.dropdown.xw(), y, False)
        self.remove = Button("Reports", "-", self.add.x() + self.add.width(), y, False)
        self.change = Button("Reports", "Edit", self.remove.xw(), y, False)

        self.add.clicked.connect(self.addOption)
        self.remove.clicked.connect(self.removeOption)
        self.change.clicked.connect(self.editOption)

        self.manual_delete = False

    def shift(self, x, y):
        self.checkbox.move(x, y)
        self.label.move(x + self.checkbox.width(), y)
        self.dropdown.move(self.label.xw(), y)
        self.add.move(self.dropdown.xw(), y)
        self.remove.move(self.add.xw(), y)
        self.change.move(self.remove.xw(), y)

    def delete(self):
        self.dropdown.options = None
        self.dropdown.deleteLater()
        self.label.deleteLater()
        self.checkbox.deleteLater()
        self.add.deleteLater()
        self.remove.deleteLater()
        self.change.deleteLater()

    def addOption(self):
        text, ok = QInputDialog(window.getScreen("Reports")).getText(window.getScreen("Reports"), "Add Option", "Sentence", QLineEdit.Normal, "")
        if ok and len(text) > 0:
            self.dropdown.options.append(text)
            self.dropdown.addItem(replace_generics(text))
            self.dropdown.setCurrentIndex(self.dropdown.count()-1)
            self.checkbox.setChecked(True)

        sentence_thread = Thread(target=self.write_sentences)
        sentence_thread.start()

    def removeOption(self):
        if self.dropdown.count() > 0:
            SentenceGroup.dialog.setComboBoxItems(self.dropdown.options)
            SentenceGroup.dialog.setWindowTitle("Remove Option")
            SentenceGroup.dialog.setLabelText("Choose an item to remove:")

            if SentenceGroup.dialog.exec():
                if SentenceGroup.dialog.textValue() in self.dropdown.options:
                    self.dropdown.options.remove(SentenceGroup.dialog.textValue())
                    self.dropdown.clear()
                    self.dropdown.addItems([replace_generics(option) for option in self.dropdown.options])
                    if self.dropdown.count() == 0:
                        self.checkbox.setChecked(False)
                    sentence_thread = Thread(target=self.write_sentences)
                    sentence_thread.start()
                else: print("Tried to invalid remove!")
        else:
            global sentences
            sentences.remove(self)
            self.delete()


            for i in range(0, len(sentences)):
                if sentences[i].index > self.index:
                    sentences[i].index -= 1
                    sentences[i].shift(sentences[i].x, sentences[i].y - 25)
                    sentences[i].label.setText("S{}:".format(i + 1))

            self.manual_delete = True
            sentence_thread = Thread(target=self.write_sentences)
            sentence_thread.start()


    def editOption(self):
        if self.dropdown.count() > 0:
            SentenceGroup.dialog.setComboBoxItems(self.dropdown.options)
            SentenceGroup.dialog.setWindowTitle("Edit Item")
            SentenceGroup.dialog.setLabelText("Choose an item to edit:")
            if SentenceGroup.dialog.exec():
                replace, ok = QInputDialog(window.getScreen("Reports")).getText(window.getScreen("Reports"), "Edit Option", "Replace '{}' with:".format(SentenceGroup.dialog.textValue()), QLineEdit.Normal, SentenceGroup.dialog.textValue())
                if ok and len(replace) > 0:
                    index = self.dropdown.options.index(SentenceGroup.dialog.textValue())
                    self.dropdown.options[index] = replace
                    self.dropdown.clear()
                    self.dropdown.addItems([replace_generics(option) for option in self.dropdown.options])
                    self.dropdown.setCurrentIndex(index)

                    sentence_thread = Thread(target=self.write_sentences)
                    sentence_thread.start()
        else:
            self.addOption()

    def write_sentences(self):
        global sheet
        global class_dropdown
        global report_sheet
        global sentence_tabs
        global all_tab_pairs

        col = index_to_column(self.index)
        sentence_tab = "Sentences " + class_dropdown.currentText().split("-")[0] #unsure why, sentence_tab() not working?
        if prefs.get_pref("is_web") and not sentence_tab in all_tab_pairs:
            raise IndexError(f"Sentence tab {sentence_tab} not in all tab pairs!")

        if self.manual_delete:
            if prefs.get_pref('is_web'):
                write_sheet(report_sheet, "", mode="COLUMNS", tab_id=all_tab_pairs[sentence_tab]['sheetId'], option={"operation":"remove", "start":self.index, "end":self.index+1})
                all_tab_pairs[sentence_tab]['gridProperties']['columnCount'] -= 1
            else:
                sheet[sentence_tab].delete_cols(self.index+1)
                sheet.save(prefs.get_pref('report_sheet'))
        else:
            buffer_list = ["", "", "", "", ""] #account for the fact that in removing, there are floating cells not part of options; have a 5-long buffer to manage that
            write_list = self.dropdown.options + buffer_list
            if prefs.get_pref('is_web'):
                write_sheet(report_sheet, [write_list], mode="COLUMNS", r="{}!{}2:{}".format(sentence_tab, col, col + str(len(write_list) + 1)))
            else:
                for i in range(0, len(write_list)):
                    sheet[sentence_tab][col+str(i+2)] = write_list[i]
                    sheet.save(prefs.get_pref('report_sheet'))

def reformat_student_dropdown():
    global student_dropdown
    global class_students

    student_dropdown.clear()
    model = student_dropdown.model()
    for i in range(len(class_students)):
        item = QStandardItem(class_students[i].first_name + " " + class_students[i].last_name)
        if not class_students[i].submitted and not class_students[i].report:
            item.setBackground(QColor(255, 0, 0))
        elif not class_students[i].submitted:
            item.setBackground(QColor(255, 255, 0))
        model.appendRow(item)

def fill_class_data(class_index=None):
    global report_sheet
    global sheet
    global class_students
    global class_dropdown
    global class_tabs
    global student_dropdown
    global report_area
    global row_offset

    student_dropdown.clear()
    class_students = []
    if class_index:
        class_name = class_tabs[class_index]
    elif class_dropdown.count() > 0:
        class_name = class_dropdown.currentText()
    else:
        class_name = None

    if class_name:
        if prefs.get_pref('is_web'):
            current_class = get_sheet(report_sheet, "{}!A1:Z1000".format(class_name)).get('values')
        else:
            current_class = [row for row in sheet[class_name].values]

        ro = row_offset #Row offset

        #Drop the headers, iterate over all student rows
        if current_class:
            assignment_headers = None
            scheme_headers = None
            grade_list = None
            if len(current_class[0]) >= 5:
                scheme_headers = [c.split("$")[-1] if len(c.split("$")) == 2 else 'IB' for c in current_class[0][4:]]
                assignment_headers = [c.split("$")[0] if len(c.split("$")) == 2 else c for c in current_class[0][4:]]
                grade_list = [""]*len(assignment_headers) #make blank list of grades, as grade # might be > assignment header

            for student in current_class[1:]:
                if grade_list:
                    for i in range(0, len(grade_list)): #loop over grade_list to fill
                        if i >= len(student[4:]): grade_list[i] = ""
                        else: grade_list[i] = student[4:][i]
                class_students.append(
                    Student(
                        student[0] if len(student) >= 1 else "",  # First Name
                        student[1] if len(student) >= 2 else "",  # Last Name
                        student[2] if len(student) >= 3 else "",  # Gender
                        student[3] if len(student) >= 4 else "",  # Report
                        {assignment: {'grade': grade, 'scheme': scheme} for assignment, grade, scheme in
                         zip(assignment_headers, grade_list, scheme_headers)} if assignment_headers else {},
                        class_name,
                        ro
                    )
                )
                ro+=1

        if prefs.get_pref('format_unfinished'):
            reformat_student_dropdown()
        else:
            student_dropdown.addItems([student.first_name + " " + student.last_name for student in class_students if student.first_name and student.last_name])
        if class_dropdown.history[-1].split("-")[0] != class_name.split("-")[0] or not first_run:
            update_sentences()
        class_dropdown.history.append(class_name)
        update_report()
        setup_grades_table()

def update_tab_order():
    global sentences
    global class_dropdown
    global student_dropdown
    global report_area

    window.getScreen("Reports").setTabOrder(class_dropdown, student_dropdown)

    if len(sentences) > 0:
        window.getScreen("Reports").setTabOrder(student_dropdown, sentences[0].dropdown)
        count = 0
        if len(sentences) > 1:
            for sentence in sentences[:-1]:
                window.getScreen("Reports").setTabOrder(sentences[count].dropdown, sentences[count+1].dropdown)
                count+=1
        window.getScreen("Reports").setTabOrder(sentences[count].dropdown, generate_button)
    else:
        window.getScreen("Reports").setTabOrder(student_dropdown, generate_button)
    window.getScreen("Reports").setTabOrder(generate_button, report_area)
    window.getScreen("Reports").setTabOrder(report_area, submit_button)

def update_sentences():
    global sentences
    global class_dropdown
    global sheet
    if len(sentence_tabs) > 0:
        for tab in sentence_tabs:
            if tab.endswith(class_dropdown.currentText().split("-")[0]):
                for elem in sentences: elem.delete()
                sentences = []
                print("Called update sentences...")
                if prefs.get_pref("is_web"):
                    current_sentences = get_sheet(report_sheet, "{}!A1:Z1000".format(sentence_tab()), "COLUMNS").get('values')
                else:
                    st = sheet[sentence_tab()]
                    current_sentences = [list(filter(None, col)) for col in st.iter_cols(values_only=True)]
                count = 0
                ro = 0
                if current_sentences is not None:
                    for entry in current_sentences:
                        if len(entry) > 1:
                            sentences.append(
                                SentenceGroup("S{}:".format(count+1), 50, 150 + 25 * count, list(filter(None, entry[1:])), ro)
                            )
                            count += 1
                        ro+=1
                populate_presets(current_sentences)
                break
    update_tab_order()

def add_sentence():
    global sentences
    global all_tab_pairs
    sentences.append(
        SentenceGroup(f"S{sentences.__len__()+1}:", 50, 125 + 25 * (sentences.__len__()+1), [], (sentences.__len__()))
    )
    cc = all_tab_pairs[sentence_tab()]['gridProperties']['columnCount']
    if cc < len(sentences):
        Thread(target=lambda: write_sheet(report_sheet, [], mode="COLUMNS", tab_id=all_tab_pairs[sentence_tab()]['sheetId'], option={"operation":"insert", "start":cc, "end":cc+1})).start()
        all_tab_pairs[sentence_tab()]['gridProperties']['columnCount']+=1
    update_tab_order()

class Preset:
    prefix = {
        "grade":"~",
        "linear":"*"
    }
    def __init__(self, text, group, index, ruleset=None, priority=None):
        self.text = text
        self.group = group
        self.index = index
        self.ruleset = ruleset
        self.priority = priority

    def __repr__(self):
        return f'Text: {self.text} | Group: {self.group} | Index: {self.index} | Ruleset: {self.ruleset} | Priority: {self.priority}'

def populate_presets(sentence_set):
    global preset_dropdown
    global preset_list
    global grade_rules
    preset_dropdown.clear()
    presets_found = []
    preset_list = {}
    grade_rules = []
    count = 1
    if sentence_set is not None:
        for elem in sentence_set:
            for cell in elem:
                #Split cell based on Preset delimiters
                cell_parts = re.split(f"[{Preset.prefix['grade']}{Preset.prefix['linear']}]", cell)
                if len(cell_parts) > 1:
                    preset_text = cell_parts[0] #Sentence comes first
                    for ps in cell_parts[1:]:
                        preset_parts = re.split("[|]", ps) #Split cell based into its components
                        if len(preset_parts) == 3: #3-part ones have a ruleset string
                            grade_rules.append(
                                Preset(preset_text, preset_parts[0], count, ruleset=preset_parts[1], priority=preset_parts[2])
                            )
                        elif len(preset_parts) == 2:
                            presets_found.append(
                                Preset(preset_text, preset_parts[0], preset_parts[1])
                            )
            count+=1

        for elem in presets_found:
            if not elem.group in preset_list:
                preset_list[elem.group] = []
            preset_list[elem.group].append(elem)

        for elem in preset_list:
            preset_list[elem] = sorted(preset_list[elem], key=lambda e: int(e.index))
            preset_dropdown.addItem(elem)

def update_report(index=None):
    global class_students
    global student_dropdown
    global report_area
    if index:
        report_area.setText(class_students[index].report)
    elif student_dropdown.count() > 0:
        report_area.setText(class_students[student_index()].report)
    else:
        report_area.setText("")
    report_area.repaint()

def student_index():
    global student_dropdown
    global class_students
    idx = student_dropdown.currentIndex()
    dropdown_items = [student_dropdown.itemText(i).strip() for i in range(student_dropdown.count())]
    if len(dropdown_items) > 0 and len(class_students) > 0:
        if not len(dropdown_items) == len(class_students):
            for i, s in enumerate(class_students):
                if dropdown_items[idx] == s.full_name():
                    return i
        else:
            return idx
    return False

def send_report():
    global class_students
    global student_dropdown

    idx = student_index()
    if idx is not False:
        class_students[idx].report = report_area.toPlainText()
        submit_thread = Thread(target=class_students[idx].submit_report)
        submit_thread.start()

    if prefs.get_pref('reformat_unfinished'):
        current_index = student_dropdown.currentIndex()
        reformat_student_dropdown()
        student_dropdown.setCurrentIndex(current_index)

def generate_report():
    global sentences
    global report_area
    global student_dropdown
    global class_students

    report_area.setText("")
    idx = student_index()
    if idx is not False:
        for sentence in sentences:
            if sentence.checkbox.isChecked():
                report_area.setText(report_area.toPlainText() + replace_generics(sentence.dropdown.options[sentence.dropdown.currentIndex()]) + " ")
    report_area.repaint()

def generate_report_from_preset():
    global report_area
    global student_dropdown
    global class_students
    global preset_list
    global preset_dropdown

    report_area.setText("")
    if preset_dropdown.count() > 0 and student_dropdown.count() > 0:
        for elem in preset_list[preset_dropdown.currentText()]:
            report_area.setText(report_area.toPlainText() + replace_generics(elem.text) + " ")
    report_area.repaint()

grades_table = Table('Grades', header=["Assignment", "Grade", "Scheme"], locked=["Assignment", "Scheme"], x=0, y=0, w=700, h=window.height(), custom_change=True)

class SheetBuilder:
    def __init__(self, options=None, tables=None):
        self.options = options if options else []
        self.tables = tables if tables else []
        self.dropdown = Dropdown("Builder", x=window.width()/2.25, y=0, options=self.options, editable=False)
        #Set of page options
        self.add_options = ["Add Page...", "Class", "Sentences", "Schemes", "Other"]

        self.sentence_headers = ["Sentences 1", "Sentences 2", "Sentences 3", "Sentences 4"]
        self.class_headers = ["First", "Last", "Gender", "Report"]

        self.add_dropdown = Dropdown("Builder", x=self.dropdown.xw(), y = self.dropdown.y(), options=self.add_options)
        self.remove_button = Button("Builder", "-", x=self.add_dropdown.xw(), y=self.dropdown.y(), focusOnTab=False)
        self.add_class_dialog = Multidialog("Builder", "Add Class", [{"name":"Class", "label":"Class", "type":"input"}, {"name":"Block", "label":"Block", "type":"input"}])

        self.veracross_button = Button("Builder", "Load from Veracross", x=window.width()/4, y=700)
        self.excel_button = Button("Builder", "Save Excel", x=self.veracross_button.xw(), y=self.veracross_button.y())
        self.sheets_button = Button("Builder", "Save Sheets", x=self.excel_button.xw(), y=self.excel_button.y())
        self.dropdown.currentIndexChanged.connect(self.change_class)

        self.excel_button.clicked.connect(self.create_excel_sheet)
        self.sheets_button.clicked.connect(self.create_google_sheet)

        self.veracross_button.clicked.connect(self.generate_from_veracross)
        self.veracross_login_dialog = Multidialog("Builder", "Input Login", [{"name":"Username", "label":"Username", "type":"input"},
                                                                             {"name":"Password", "label":"Password", "type":"input", "settings":{"mode":"password"}}])

        self.add_scheme_dialog = Multidialog("Builder", "Create Scheme", [{"name":"Name", "label":"Name", "type":"input"},
                                                                          {"name":"Scale", "label":"Scale", "type":"dropdown", "data":["Increasing", "Decreasing"]},
                                                                          {"name":"Type", "label":"Element Type", "type":"dropdown", "data":["Integer", "Number", "Letter", "List"]},
                                                                          {"name":"Lower Bound", "label":"Lower Bound", "type":"input", "settings":{"conditional_show":{
                                                                              "element":"Type", "not_state":["List"]
                                                                          }}},
                                                                          {"name":"Upper Bound", "label":"Upper Bound", "type":"input", "settings":{"conditional_show":{
                                                                              "element":"Type", "not_state":["List"]
                                                                          }}},
                                                                          {"name":"Pass Bound", "label":"Pass Bound", "type":"input", "settings":{"conditional_show":{
                                                                              "element":"Type", "not_state":["List"]
                                                                          }}},
                                                                          {"name":"Options", "label":"Options (,)", "type":"input", "settings":{"conditional_show":{
                                                                              "element":"Type", "state":["List"]
                                                                          }}}])

        self.add_dropdown.currentIndexChanged.connect(self.add_option)
        # self.remove_button.clicked.connect(self.remove_option)

    def create_excel_sheet(self):
        global report_sheet
        file = QFileDialog.getSaveFileName(window.getScreen("Builder"), 'Dialog Title', os.path.join(os.getcwd(), "data"), '(*.xlsx)')
        if len(file[0]) > 0:
            excel_sheet = openpyxl.Workbook()
            for i, tab in enumerate(self.options):
                excel_sheet.create_sheet(title=tab)
                excel_sheet[tab].append(self.tables[i].header)
                if self.tables[i].data:
                    for j, row in enumerate(self.tables[i].data):
                        excel_sheet[tab].append(row)
            excel_sheet.remove(excel_sheet[excel_sheet.sheetnames[0]]) #Remove default tab
            excel_sheet.save(file[0])
            prefs.update_pref("is_web", False)
            prefs.update_pref("report_sheet", file[0])
            report_sheet = file[0]
            setup_existing()

    def create_google_sheet(self):
        name, ok = QInputDialog(window.getScreen("Builder")).getText(window.getScreen("Builder"), "Name Google Sheet", "Sheet Title:", QLineEdit.Normal, "")
        if ok:
            if len(name) > 0:
                req_data = [{"title":tab, "data":[self.tables[i].header]+self.tables[i].data} for i, tab in enumerate(self.options)]
                report_sheet = make_report_sheet(name, req_data)
                prefs.update_pref("is_web", True)
                prefs.update_pref("report_sheet", report_sheet)
                setup_existing()

    def generate_from_veracross(self):
        if self.veracross_login_dialog.exec():
            username = self.veracross_login_dialog.elements['Username']['object'].text().strip()
            password = self.veracross_login_dialog.elements['Password']['object'].text()
            if len(username) > 0 and len(password) > 0:
                class_json = get_class_json(username=username, password=password)
                if not class_json:
                    print("Invalid login!", username, password)
                    return
                self.dropdown.clear()
                self.options = []
                for t in self.tables: t.deleteLater()
                self.tables = []
                for c in class_json:
                    class_name = class_json[c]['name']
                    student_set = [[s['preferred_name'], s['last_name'], s['gender'], ""] for s in class_json[c]['students']]
                    self.tables.append(Table("Builder", header=self.class_headers, x=0, y=30, w=window.width(), h=675, shown=False, accepted={"Gender":{"whitelist":["M", "F", "T"]}}))
                    self.tables[-1].updateTable(student_set)
                    self.options.append(class_name)
                self.dropdown.addItems(self.options)
                self.tables[-1].show()

    def change_class(self):
        if len(self.tables) > 0:
            for t in self.tables: t.hide()
            self.tables[self.dropdown.currentIndex()].show()

    #naive add/remove
    def add_option(self, index):
        if index != 0:
            if self.add_dropdown.currentText() == "Class":
                if self.add_class_dialog.exec():
                    class_name = self.add_class_dialog.elements['Class']['object'].text().replace("-", " ").strip()
                    block_name = self.add_class_dialog.elements['Block']['object'].text().replace("-", " ").strip()
                    if len(class_name) > 0 and len(block_name) > 0:
                        tab_name = class_name+"-"+block_name
                        sentences_name = "Sentences " + class_name
                        self.add_sentence_tab(sentences_name)
                        self.options.append(tab_name)
                        self.dropdown.addItem(tab_name)
                        self.tables.append(Table("Builder", header=self.class_headers, x=0, y=30, w=window.width(), h=675, shown=True, accepted={"Gender":{"whitelist":["M", "F", "T"]}}))
                        self.tables[-1].updateTable([[""]*4]*10)
                        self.dropdown.setCurrentIndex(self.dropdown.count() - 1)
                    self.add_class_dialog.refresh()
            elif self.add_dropdown.currentText() == "Sentences":
                course_name, ok = QInputDialog(window.getScreen("Builder")).getText(window.getScreen("Builder"), "Add Sentence Tab", "Course Name:", QLineEdit.Normal, "")
                if ok: self.add_sentence_tab("Sentences " + course_name.replace("-", ""), True)
            elif self.add_dropdown.currentText() == "Schemes":
                if self.add_scheme_dialog.exec():
                    scheme_name = self.add_scheme_dialog.elements["Name"]['object'].text().strip()
                    if len(scheme_name) > 0:
                        scheme_scale = [gs for gs in GradeScale][self.add_scheme_dialog.elements["Scale"]['object'].currentIndex()]
                        scheme_type = [gt for gt in GradeType][self.add_scheme_dialog.elements["Type"]['object'].currentIndex()]

                        scheme_pb = self.add_scheme_dialog.elements["Pass Bound"]['object'].text().strip()
                        scheme_lb = self.add_scheme_dialog.elements["Lower Bound"]['object'].text().strip()
                        scheme_ub = self.add_scheme_dialog.elements["Upper Bound"]['object'].text().strip()
                        scheme_options = [o.strip() for o in self.add_scheme_dialog.elements['Options']['object'].text().split(",")]

                        idx_s = self.dropdown.findText("Grade Schemes")
                        if idx_s == -1: #scheme tab does not exist
                            self.options.append("Grade Schemes")
                            self.dropdown.addItem("Grade Schemes")
                            self.tables.append(Table("Builder", header=[scheme_name], x=0, y=30, w=window.width(), h=675, shown=True))
                            if scheme_type == GradeType.SET: self.tables[-1].updateTable([[o] for o in scheme_options])
                            else: self.tables[-1].updateTable([[scheme_ub], [scheme_lb]])
                            self.dropdown.setCurrentIndex(self.dropdown.count() - 1)
                        else:
                            self.tables[idx_s].insertColumn(self.tables[idx_s].columnCount())
                            self.tables[idx_s].header += [scheme_name]
                            self.tables[idx_s].setHorizontalHeaderLabels(self.tables[idx_s].header)
                            if scheme_type == GradeType.SET:
                                for i, item in enumerate(scheme_options):
                                    self.tables[idx_s].setItem(i, self.tables[idx_s].columnCount() - 1, QTableWidgetItem(item))
                            else:
                                self.tables[idx_s].setItem(0, self.tables[idx_s].columnCount() - 1, QTableWidgetItem(scheme_ub))
                                self.tables[idx_s].setItem(1, self.tables[idx_s].columnCount() - 1, QTableWidgetItem(scheme_lb))
                        self.add_scheme_dialog.refresh()

        self.add_dropdown.setCurrentIndex(0)

    def add_sentence_tab(self, sn, show=False):
        if len(sn) > 0 and self.dropdown.findText(sn) == -1:
            self.options.append(sn)
            self.dropdown.addItem(sn)
            self.tables.append(Table("Builder", header=self.sentence_headers, x=0, y=30, w=window.width(), h=675, shown=show))
            self.tables[-1].updateTable([[""] * 4] * 10)
            if show: self.dropdown.setCurrentIndex(self.dropdown.count() - 1)

sheet_builder = SheetBuilder()

def grade_cell_changed(item):
    global class_students
    global student_dropdown

    if grades_table.oldCell != (item.row(), item.column()): grades_table.oldText = ""
    if class_students is not None and student_dropdown.count() > 0:
        current_student = class_students[student_index()]
        col = item.column()
        row = item.row()

        if grades_table.horizontalHeaderItem(col) is not None and grades_table.horizontalHeaderItem(col).text() == "Grade":
            scheme = grades_table.item(row, col + 1)
            assignment = grades_table.item(row, col - 1)
            if scheme is not None and not GradeSet(scheme.text()).is_valid(item.text()):
                item.setText(grades_table.oldText)
            elif scheme is not None:
                current_student.grades[assignment.text()]['grade'] = item.text()

                grades_thread = Thread(target=current_student.write_grades)
                grades_thread.start()
            grades_table.oldText = item.text()
        #If setCurrentCell is called in keyPressEvent, the validation step occurs AFTER backup step, meaning data is overwritten incorrectly
        if grades_table.enterPressed:
            grades_table.setCurrentCell(row + 1, col)
            grades_table.enterPressed = False

    grades_table.resizeRowsToContents()
    grades_table.resizeColumnsToContents()


grades_table.itemChanged.connect(grade_cell_changed)

def generate_report_from_grades():
    global class_students
    global grade_rules
    global class_dropdown

    report_area.setText("")
    idx = student_index()
    if idx is not False:
        current_student = class_students[student_index()]
        chosen_options = {}
        for elem in grade_rules:
            tk = GradeSet.tokenize(elem.ruleset)
            use_scheme = 'IB'
            if len(tk) == 3:
                if tk[0] in current_student.grades: use_scheme = current_student.grades[tk[0]]['scheme']
                elif tk[2] in current_student.grades: use_scheme = current_student.grades[tk[2]]['scheme']

                if GradeSet(use_scheme).evaluate(current_student.grades, elem.ruleset):
                    if elem.index in chosen_options:
                        #greater than because number is inverted
                        if chosen_options[elem.index].priority > elem.priority:
                            chosen_options[elem.index] = elem
                    else:
                        chosen_options[elem.index] = elem
        for option in chosen_options:
            report_area.setText(report_area.toPlainText() + replace_generics(chosen_options[option].text) + " ")
    report_area.repaint()

def setup_grades_table():
    global class_students
    global student_dropdown
    global grades_table

    idx = student_index()
    if idx is not False:
        current_student = class_students[idx]
        grades_table.updateTable([[g, current_student.grades[g]['grade'], current_student.grades[g]['scheme']] for g in current_student.grades])
    else:
        grades_table.updateTable()

def update_student(index=None):
    global sentences

    update_report(index)
    setup_grades_table()
    for sentence in sentences:
        sentence.dropdown.clear()
        sentence.dropdown.addItems([replace_generics(formatted) for formatted in sentence.dropdown.options])

def local_save_report():
    global class_students
    global student_dropdown
    idx = student_index()
    if idx is not False: class_students[idx].report = report_area.toPlainText()

first_run = True
def setup():
    global sheet
    global all_tab_pairs
    global all_tabs
    global class_tabs
    global sentence_tabs
    global grade_scheme_tabs
    global class_students
    global preset_list
    global grade_rules
    global first_run

    if prefs.get_pref('is_web'):
        all_tab_pairs = {tab['properties']['title']:tab['properties'] for tab in sheet.get('sheets')}
        all_tabs = [tab for tab in all_tab_pairs]
    else:
        all_tabs = sheet.sheetnames

    grade_scheme_tabs = [tab for tab in all_tabs if tab.startswith("Grade Scheme")]
    user_template_tabs = [tab for tab in all_tabs if tab.startswith("User Template")]
    class_tabs = [tab for tab in all_tabs if len(tab.split("-")) == 2]
    sentence_tabs = [tab for tab in all_tabs if tab.startswith("Sentences")]

    class_students = []
    preset_list = {}
    grade_rules = []

    window.switchScreen("Reports")
    open_reports_from_setup_button.show()

    if not first_run:
        class_dropdown.currentIndexChanged.disconnect()
        student_dropdown.currentIndexChanged.disconnect()
    class_dropdown.clear()
    class_dropdown.addItems([tab for tab in class_tabs])
    load_grades(grade_scheme_tabs)
    load_templates(user_template_tabs)
    fill_class_data()
    class_dropdown.currentIndexChanged.connect(fill_class_data)
    student_dropdown.currentIndexChanged.connect(update_student)
    toggle_advanced([preset_dropdown, preset_button, grade_button, open_setup_from_report_button], force=prefs.get_pref("advanced"))
    first_run = False

def setup_existing():
    global sheet
    global report_sheet

    valid, sheet = validate_sheet(prefs.get_pref("report_sheet"))
    if valid:
        report_sheet = prefs.get_pref("report_sheet")
        setup()


def setup_sheet_from_file():
    global sheet
    global report_sheet

    file = file_select("Excel Sheet", excel_extensions)
    if file:
        try:
            sheet = openpyxl.load_workbook(file)
            prefs.update_pref("is_web", False)
            prefs.update_pref("report_sheet", file)
            report_sheet = prefs.get_pref("report_sheet")
            setup()
        except openpyxl.utils.exceptions.InvalidFileException:
            return StatusCode.INVALID_PATH

class StatusCode(Enum):
    NONE = (0, "")
    SUCCESS = (1, "Success!")
    INVALID_PATH = (2, "Invalid filepath!")
    PERMISSION_FAILURE = (3, "Invalid permissions to access document!")
    INVALID_URL = (3, "Invalid path URL!")
    INVALID_ID = (4, "Invalid path ID!")
    INTERNET_FAILURE = (5, "No internet connection!")


    def __init__(self, code, text):
        self.code = code
        self.text = text

def validate_sheet(report):
    global report_sheet

    if os.path.isfile(report) and report.endswith(tuple(excel_extensions)):
        try:
            workbook = openpyxl.load_workbook(report)
            prefs.update_pref("is_web", False)
            prefs.update_pref("report_sheet", report)
            return True, workbook
        except openpyxl.utils.exceptions.InvalidFileException:
            return False, StatusCode.INVALID_PATH
    elif not os.path.isfile(report):
        check_is = ["d", "u", "1", "https:", "spreadsheets", "http:", "docs.google.com", '']
        check_starts = ["edit#"]

        report = [part for part in report.split('/') if not part in check_is and not part.startswith(*check_starts)]
        if len(report) != 1:
            return False, StatusCode.INVALID_ID
        else:
            try:
                sheet_data = get_sheet(report[0])
                report_sheet = report[0]
                prefs.update_pref("report_sheet", report[0])
                prefs.update_pref("is_web", True)
                return True, sheet_data
            except googleapiclient.errors.HttpError as e:
                error_data = json.loads(e.content)['error']
                if error_data['status'] == "PERMISSION_DENIED": return False, StatusCode.PERMISSION_FAILURE
                elif error_data['status'] == "NOT_FOUND": return False, StatusCode.INVALID_URL
            except httplib2.ServerNotFoundError:
                return False, StatusCode.INTERNET_FAILURE
    else:
        return False, StatusCode.INVALID_PATH


def setup_sheet_from_dialog(report=None):
    global report_sheet
    global sheet

    prompt_text = "Input Google Spreadsheet link or Excel path:"
    response = StatusCode.NONE
    valid = False
    while not valid:
        if not report:
            report, ok = QInputDialog(window.getScreen("Reports")).getText(window.getScreen("Reports"), "Report Sheet", (response.text+" "+prompt_text).strip(), QLineEdit.Normal, "")
            if not ok:
                sheet = False
                return
        valid, response = validate_sheet(report)
        report = None
    sheet = response
    setup()

def launch_sheet():
    if prefs.get_pref("is_web"): webbrowser.open_new(f"https://docs.google.com/spreadsheets/d/{prefs.get_pref('report_sheet')}")
    else:
        if is_macos():
            os.system(f"open {prefs.get_pref('report_sheet')}")
        else:
            os.startfile(prefs.get_pref('report_sheet'))

def open_sheet():
    ost = Thread(target=launch_sheet)
    ost.start()

def copy_report():
    global report_area
    text = report_area.toPlainText()
    cb = app.clipboard()
    cb.clear(mode=cb.Clipboard)
    cb.setText(text if len(text) > 0 else "")

def load_templates(template_tabs):
    global user_templates
    user_templates = {}
    if template_tabs:
        for tab in template_tabs:
            if prefs.get_pref("is_web"):
                uts = get_sheet(prefs.get_pref('report_sheet'), "{}!A2:C1000".format(tab), mode='ROWS').get('values')
            else:
                uts = [[r if r is not None else "" for r in row] for row in sheet[tab].values]
            if uts:
                for ut in uts:
                    if len(ut) >= 2:
                        if len(ut[0].strip()) == 0 or len(ut[1].strip()) == 0: continue
                        else: user_templates[ut[0]] = {"value":ut[1]}

                        if len(ut) >= 3 and len(ut[2].strip()) > 0: user_templates[ut[0]]["class"] = ut[2]


def toggle_advanced(objs, force=None):
    if len(objs) > 0:
        for o in objs:
            o.setVisible(not o.isVisible() if force is None else force)
            o.repaint()
        toggle_advanced_button.setText("Hide Advanced" if objs[0].isVisible() else "Show Advanced")
        toggle_advanced_button.repaint() #repaint calls to fix a big on some laptops where this doesn't hide/show
        prefs.update_pref("advanced", objs[0].isVisible())

def display_help():
    #todo: implement help menu lmao
    pass

load_sheet_file_button = Button("Setup", "Load Sheet (File)", window.width() / 3, window.height() / 3, False)
load_sheet_url_button = Button("Setup", "Load Sheet (URL)", load_sheet_file_button.xw(), window.height() / 3, False)

create_report_sheet_button = Button("Setup", "Create Reports Sheet", window.width() / 3, load_sheet_file_button.y() + load_sheet_file_button.height(), False)
open_setup_from_builder_button = Button("Builder", "Back", 0, 0, False)

load_sheet_file_button.clicked.connect(setup_sheet_from_file)
load_sheet_url_button.clicked.connect(setup_sheet_from_dialog)
create_report_sheet_button.clicked.connect(lambda: window.switchScreen("Builder"))
open_setup_from_builder_button.clicked.connect(lambda: window.switchScreen("Setup"))

report_area.textChanged.connect(local_save_report)
submit_button.clicked.connect(send_report)
generate_button.clicked.connect(generate_report)
preset_button.clicked.connect(generate_report_from_preset)
grade_button.clicked.connect(generate_report_from_grades)
reload_all_button.clicked.connect(setup_existing)
add_sentence_button.clicked.connect(add_sentence)
launch_report_sheet_button.clicked.connect(open_sheet)
copy_button.clicked.connect(copy_report)
toggle_advanced_button.clicked.connect(lambda: toggle_advanced([preset_dropdown,preset_button,grade_button,open_setup_from_report_button]))
# Attempting to clean up UI, might bring back
# reload_sentences_button.clicked.connect(update_sentences)
# reload_schemes_button.clicked.connect(lambda: load_grades(grade_scheme_tabs))

help_button.clicked.connect(display_help)

color_button = ColorButton("Preferences", 0, 0, 100, 100, "#ffffff")

if __name__ == "__main__":
    if len(report_sheet) > 0: setup_existing()
    app.exec()


